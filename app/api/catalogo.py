import io
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import get_db
from app.security import get_current_user
from app.models.catalogo import (
    CuentaPlanificacion,
    Gestor,
    ItemPlanificacion,
    PlanillaTemplate,
    PlanPresupuestario,
    RelacionItemCuenta,
    TipoMovimiento,
)
from app.schemas.catalogo import (
    CuentaOut,
    GestorOut,
    ItemOut,
    PlanillaTemplateOut,
    PlanOut,
    TipoMovimientoOut,
)

# Datos maestros: visibles para cualquier usuario autenticado (no se filtran
# por VP — todos los planificadores necesitan el mismo catálogo). La auth se
# aplica una vez a nivel router para cerrar todo el módulo en bloque.
router = APIRouter(
    prefix="/catalogo",
    tags=["catalogo"],
    dependencies=[Depends(get_current_user)],
)


# ============================================================
# Mapeo planilla → cuentas (espejo del PLANILLA_COMPONENTES del frontend)
# Esta es la fuente canónica para reportes/export. El frontend lo replica
# para evaluar las fórmulas en vivo.
# ============================================================
MAPA_PLANILLA_CUENTAS: dict[str, list[dict[str, str]]] = {
    "PL-MISIONES-SERV": [
        {"cuenta_codigo": "5.4.1.01", "concepto": "Pasajes",   "formula": "cant_viajes × tarifa_pasaje(destino)"},
        {"cuenta_codigo": "5.4.1.02", "concepto": "Viáticos",  "formula": "cant_viajes × días × tarifa_viatico(destino)"},
        {"cuenta_codigo": "5.4.1.03", "concepto": "Hospedaje", "formula": "cant_viajes × días × tarifa_hospedaje(destino)"},
    ],
    "PL-MISIONES-CONS": [
        {"cuenta_codigo": "5.3.1.02", "concepto": "Pasajes consultores",   "formula": "cant_viajes × tarifa_pasaje(destino)"},
        {"cuenta_codigo": "5.3.1.03", "concepto": "Viáticos consultores",  "formula": "cant_viajes × días × tarifa_viatico(destino)"},
        {"cuenta_codigo": "5.3.1.04", "concepto": "Hospedaje consultores", "formula": "cant_viajes × días × tarifa_hospedaje(destino)"},
    ],
    "PL-CONSULTORES": [
        {"cuenta_codigo": "5.3.1.01", "concepto": "Honorarios", "formula": "valor_hora × horas_mes × meses"},
    ],
    "PL-REUNIONES-EVENTOS": [
        {"cuenta_codigo": "5.6.1.03", "concepto": "Reuniones y Eventos", "formula": "monto_total (directo)"},
    ],
    "PL-SERVICIOS-LIC": [
        {"cuenta_codigo": "5.6.5.02", "concepto": "TI Equipos y Aplicaciones",   "formula": "monto_total (directo)"},
        {"cuenta_codigo": "5.6.5.03", "concepto": "TI Servicios Mantenimiento", "formula": "monto_total (directo)"},
    ],
    "PL-SALARIOS-BENEF": [
        {"cuenta_codigo": "5.2.*", "concepto": "Salarios y Beneficios", "formula": "monto_total (directo)"},
    ],
    "PL-GASTOS-ADMIN": [
        {"cuenta_codigo": "5.6.1.*", "concepto": "Comunicación", "formula": "monto_total (directo)"},
        {"cuenta_codigo": "5.6.2.*", "concepto": "Reclutamiento y otros gastos de personal", "formula": "monto_total (directo)"},
        {"cuenta_codigo": "5.6.3.*", "concepto": "Servicios al edificio", "formula": "monto_total (directo)"},
        {"cuenta_codigo": "5.6.4.*", "concepto": "Gastos de Representación", "formula": "monto_total (directo)"},
    ],
}


@router.get("/objetivos")
async def list_objetivos(db: AsyncSession = Depends(get_db)) -> list[dict]:
    """Objetivos estratégicos institucionales — usado por la columna Objetivos
    de las planillas para vincular cada esfuerzo monetario con la estrategia."""
    rows = (await db.execute(
        text(
            """SELECT id, codigo, nombre, orden, activo
                 FROM catalogo.objetivo_estrategico
                WHERE activo = true
                ORDER BY orden, codigo"""
        )
    )).mappings().all()
    return [dict(r) for r in rows]


@router.get("/items", response_model=list[ItemOut])
async def list_items(
    path: str | None = Query(None, description="ltree path filter (ej. 'n05.*' o lquery)"),
    imputable: bool | None = None,
    tipo: str | None = Query(None, description="gastos | inversiones_capital | salarios"),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(ItemPlanificacion).order_by(ItemPlanificacion.codigo)
    if path:
        # ltree: <@ con ancestor; ~ con lquery (acepta wildcards). Detectamos por '*'.
        op = "~" if "*" in path else "<@"
        stmt = stmt.where(text(f"path {op} CAST(:p AS lquery)" if op == "~" else f"path {op} CAST(:p AS ltree)").bindparams(p=path))
    if imputable is not None:
        stmt = stmt.where(ItemPlanificacion.imputable == imputable)
    if tipo:
        stmt = stmt.where(ItemPlanificacion.tipo_presupuesto == tipo)
    return (await db.execute(stmt)).scalars().all()


@router.get("/cuentas", response_model=list[CuentaOut])
async def list_cuentas(
    path: str | None = Query(None, description="ltree path filter (ej. 'n5.n4.*' o lquery)"),
    imputable: bool | None = None,
    modalidad: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    stmt = select(CuentaPlanificacion).order_by(CuentaPlanificacion.codigo)
    if path:
        op = "~" if "*" in path else "<@"
        stmt = stmt.where(text(f"path {op} CAST(:p AS lquery)" if op == "~" else f"path {op} CAST(:p AS ltree)").bindparams(p=path))
    if imputable is not None:
        stmt = stmt.where(CuentaPlanificacion.imputable == imputable)
    if modalidad:
        stmt = stmt.where(CuentaPlanificacion.modalidad_default == modalidad)
    return (await db.execute(stmt)).scalars().all()


@router.get("/gestores", response_model=list[GestorOut])
async def list_gestores(db: AsyncSession = Depends(get_db)):
    stmt = select(Gestor).order_by(Gestor.nombre)
    return (await db.execute(stmt)).scalars().all()


@router.get("/planes", response_model=list[PlanOut])
async def list_planes(db: AsyncSession = Depends(get_db)):
    stmt = select(PlanPresupuestario).order_by(PlanPresupuestario.codigo)
    return (await db.execute(stmt)).scalars().all()


@router.get("/tipos-movimiento", response_model=list[TipoMovimientoOut])
async def list_tipos_movimiento(db: AsyncSession = Depends(get_db)):
    stmt = select(TipoMovimiento).order_by(TipoMovimiento.categoria, TipoMovimiento.k2b_codigo)
    return (await db.execute(stmt)).scalars().all()


@router.get("/items/{item_id}/cuentas-validas", response_model=list[CuentaOut])
async def cuentas_validas_para_item(item_id: int, db: AsyncSession = Depends(get_db)):
    """Devuelve las cuentas que el item puede imputar (matriz relacion_item_cuenta)."""
    stmt = (
        select(CuentaPlanificacion)
        .join(RelacionItemCuenta, RelacionItemCuenta.cuenta_id == CuentaPlanificacion.id)
        .where(RelacionItemCuenta.item_id == item_id)
        .order_by(CuentaPlanificacion.codigo)
    )
    return (await db.execute(stmt)).scalars().all()


@router.get("/relacion-item-cuenta")
async def relacion_item_cuenta(db: AsyncSession = Depends(get_db)) -> list[dict[str, str]]:
    """Matriz completa item ↔ cuenta del Excel `relaciones.xlsx`.

    Es la fuente de verdad para validar qué combinación de unidad y cuenta es válida.
    El frontend la usa para filtrar el dropdown 'Unidad' según las cuentas de la planilla activa.
    """
    rows = (await db.execute(
        text("""
            SELECT i.codigo AS item_codigo, c.codigo AS cuenta_codigo
            FROM catalogo.relacion_item_cuenta ric
            JOIN catalogo.item_planificacion i ON i.id = ric.item_id
            JOIN catalogo.cuenta_planificacion c ON c.id = ric.cuenta_id
            ORDER BY i.codigo, c.codigo
        """)
    )).mappings().all()
    return [{"item_codigo": r["item_codigo"], "cuenta_codigo": r["cuenta_codigo"]} for r in rows]


@router.get("/planillas-templates", response_model=list[PlanillaTemplateOut])
async def list_planillas_templates(db: AsyncSession = Depends(get_db)):
    stmt = (
        select(PlanillaTemplate)
        .where(PlanillaTemplate.estado == "activo")
        .order_by(PlanillaTemplate.orden)
    )
    return (await db.execute(stmt)).scalars().all()


@router.get("/planillas-templates/{codigo}", response_model=PlanillaTemplateOut)
async def get_planilla_template(codigo: str, db: AsyncSession = Depends(get_db)):
    stmt = select(PlanillaTemplate).where(PlanillaTemplate.codigo == codigo)
    template = (await db.execute(stmt)).scalar_one_or_none()
    if template is None:
        raise HTTPException(404, f"Planilla template '{codigo}' no encontrada")
    return template


# ============================================================
# Mapa de relaciones — vista integrada planilla → cuentas → K2B
# ============================================================
async def _construir_mapa(db: AsyncSession) -> list[dict[str, Any]]:
    """Construye el árbol planilla → componentes → cuentas, anotado con k2b_cuenta_id.

    Cada elemento del array es una planilla con sus componentes resueltos:
      [{
        planilla_codigo, planilla_nombre, plan_codigo, modalidad,
        componentes: [{cuenta_codigo, cuenta_descripcion, concepto, formula, k2b_cuenta_id, imputable}, ...]
      }, ...]
    """
    templates = (
        await db.execute(
            select(PlanillaTemplate).where(PlanillaTemplate.estado == "activo").order_by(PlanillaTemplate.orden)
        )
    ).scalars().all()

    # Pre-cargar todas las cuentas para resolver lookups por código (incluyendo wildcards)
    cuentas_rows = (await db.execute(select(CuentaPlanificacion))).scalars().all()
    cuentas_by_code = {c.codigo: c for c in cuentas_rows}

    def expand_codigo(codigo: str) -> list[CuentaPlanificacion]:
        """Acepta '5.4.1.01' o '5.6.*' y devuelve la lista de cuentas que matchean."""
        if codigo.endswith(".*"):
            prefijo = codigo[:-2]  # quita '.*'
            return sorted(
                [c for c in cuentas_rows if c.codigo.startswith(prefijo + ".")],
                key=lambda x: x.codigo,
            )
        c = cuentas_by_code.get(codigo)
        return [c] if c else []

    out: list[dict[str, Any]] = []
    for t in templates:
        comps_def = MAPA_PLANILLA_CUENTAS.get(t.codigo, [])
        componentes: list[dict[str, Any]] = []
        for cd in comps_def:
            for c in expand_codigo(cd["cuenta_codigo"]):
                componentes.append({
                    "cuenta_codigo": c.codigo,
                    "cuenta_descripcion": c.descripcion,
                    "nivel": c.nivel,
                    "imputable": c.imputable,
                    "k2b_cuenta_id": c.k2b_cuenta_id,
                    "modalidad_default": c.modalidad_default,
                    "concepto": cd["concepto"],
                    "formula": cd["formula"],
                })
        out.append({
            "planilla_codigo": t.codigo,
            "planilla_nombre": t.nombre,
            "planilla_descripcion": t.descripcion,
            "plan_codigo": (t.scope_filter.get("plan_codigo") or ["PRESUPDEGASTOS"])[0],
            "modalidad": t.modalidad_permitida,
            "formula_default_codigo": t.formula_default_codigo,
            "scope_filter": t.scope_filter,
            "componentes": componentes,
        })
    return out


@router.get("/mapa-relaciones")
async def mapa_relaciones(db: AsyncSession = Depends(get_db)) -> dict[str, Any]:
    """Devuelve el árbol planilla → componentes/cuentas para visualización.

    Incluye:
      - jerarquía de items (códigos dot-notation) con su k2b_item_id
      - mapeo planilla → cuentas concretas → k2b_cuenta_id (para export a K2B)
      - lista de planes presupuestarios
      - lista de tipos de movimiento K2B (ejecución)
    """
    planillas = await _construir_mapa(db)

    # Items niveles 1-3 ordenados, con su gestor asociado (de gestor_item).
    items = (
        await db.execute(
            text("""
                SELECT
                  i.codigo,
                  i.descripcion,
                  i.nivel,
                  i.parent_id,
                  i.imputable,
                  i.tipo_presupuesto,
                  i.k2b_item_id,
                  -- Gestor canónico: el primero asociado al item (es 1-a-1 según relaciones.xlsx).
                  -- COALESCE busca también heredado del padre (las áreas 02.03.* heredan el gestor de 02.03).
                  COALESCE(
                    (SELECT g.nombre FROM catalogo.gestor_item gi
                       JOIN catalogo.gestor g ON g.id = gi.gestor_id
                       WHERE gi.item_id = i.id LIMIT 1),
                    (SELECT g.nombre FROM catalogo.gestor_item gi
                       JOIN catalogo.gestor g ON g.id = gi.gestor_id
                       WHERE gi.item_id = i.parent_id LIMIT 1)
                  ) AS gestor_nombre
                FROM catalogo.item_planificacion i
                WHERE i.codigo ~ '^[0-9]{2}(\\.[0-9]{2}){0,2}$'
                ORDER BY i.codigo
            """)
        )
    ).mappings().all()

    planes = (
        await db.execute(text("SELECT codigo, nombre, tipo, k2b_plan_prefix FROM catalogo.plan_presupuestario ORDER BY codigo"))
    ).mappings().all()

    tipos_mov = (
        await db.execute(
            text("SELECT k2b_codigo, nombre, categoria, signo FROM catalogo.tipo_movimiento ORDER BY categoria, k2b_codigo")
        )
    ).mappings().all()

    return {
        "planillas": planillas,
        "items": [dict(r) for r in items],
        "planes": [dict(r) for r in planes],
        "tipos_movimiento": [dict(r) for r in tipos_mov],
    }


@router.get("/mapa-relaciones.xlsx")
async def mapa_relaciones_xlsx(db: AsyncSession = Depends(get_db)) -> StreamingResponse:
    """Export Excel del mapa de relaciones: planillas, cuentas, items, planes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    data = await mapa_relaciones(db)
    wb = Workbook()

    # Estilos
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill("solid", fgColor="2A2D33")
    h_align = Alignment(horizontal="left", vertical="center")
    sub_font = Font(bold=True, size=10, color="4A4D54")

    def style_header(ws, n_cols: int) -> None:
        for col in range(1, n_cols + 1):
            c = ws.cell(row=1, column=col)
            c.font = h_font
            c.fill = h_fill
            c.alignment = h_align

    # Hoja 1: Planillas → Cuentas (esto es el corazón del export)
    ws1 = wb.active
    ws1.title = "Planillas → Cuentas"
    ws1.append([
        "Planilla código", "Planilla nombre", "Plan presupuesto", "Modalidad",
        "Componente / Concepto", "Cuenta código", "Cuenta descripción",
        "K2B cuenta id", "Imputable", "Fórmula",
    ])
    style_header(ws1, 10)
    for p in data["planillas"]:
        if p["componentes"]:
            for comp in p["componentes"]:
                ws1.append([
                    p["planilla_codigo"], p["planilla_nombre"],
                    p["plan_codigo"], p["modalidad"],
                    comp["concepto"], comp["cuenta_codigo"], comp["cuenta_descripcion"],
                    comp.get("k2b_cuenta_id") or "",
                    "Sí" if comp["imputable"] else "No",
                    comp["formula"],
                ])
        else:
            ws1.append([
                p["planilla_codigo"], p["planilla_nombre"],
                p["plan_codigo"], p["modalidad"],
                "(sin componentes definidos)", "", "", "", "", "",
            ])
    for i, w in enumerate([18, 32, 18, 14, 30, 14, 36, 14, 10, 40], start=1):
        ws1.column_dimensions[chr(64 + i)].width = w

    # Hoja 2: Items de planificación (jerarquía) + Gestor canónico
    ws2 = wb.create_sheet("Items de Planificación")
    ws2.append([
        "Código", "Descripción", "Nivel", "Gestor", "Imputable", "Tipo presupuesto", "K2B item id",
    ])
    style_header(ws2, 7)
    for it in data["items"]:
        sangria = "  " * (int(it["nivel"]) - 1)
        ws2.append([
            sangria + str(it["codigo"]),
            it["descripcion"],
            it["nivel"],
            it.get("gestor_nombre") or "",
            "Sí" if it["imputable"] else "No",
            it["tipo_presupuesto"],
            it.get("k2b_item_id") or "",
        ])
    for i, w in enumerate([20, 50, 8, 42, 10, 18, 14], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    # Hoja 3: Planes presupuestarios K2B
    ws3 = wb.create_sheet("Planes K2B")
    ws3.append(["Código", "Nombre", "Tipo", "K2B prefix"])
    style_header(ws3, 4)
    for pl in data["planes"]:
        ws3.append([pl["codigo"], pl["nombre"], pl["tipo"], pl.get("k2b_plan_prefix") or ""])
    for i, w in enumerate([20, 40, 14, 24], start=1):
        ws3.column_dimensions[chr(64 + i)].width = w

    # Hoja 4: Tipos de movimiento K2B (para ejecución)
    ws4 = wb.create_sheet("Tipos Movimiento K2B")
    ws4.append(["K2B código", "Nombre", "Categoría", "Signo"])
    style_header(ws4, 4)
    for t in data["tipos_movimiento"]:
        ws4.append([t["k2b_codigo"], t["nombre"], t["categoria"], t["signo"]])
    for i, w in enumerate([26, 36, 16, 8], start=1):
        ws4.column_dimensions[chr(64 + i)].width = w

    # Hoja 5: Flujo end-to-end (resumen narrativo)
    ws5 = wb.create_sheet("Flujo end-to-end")
    ws5.append(["#", "Etapa", "Qué pasa", "Tabla / objeto"])
    style_header(ws5, 4)
    flujo = [
        (1, "Carga en planilla",       "El usuario completa una fila en una planilla del SolicitudEditor",
         "planificacion.linea_solicitud"),
        (2, "Splitter de cuentas",     "La fila se descompone en N líneas, una por cuenta (Pasajes 5.4.1.01, Viáticos 5.4.1.02, etc.)",
         "PLANILLA_COMPONENTES (frontend) ↔ MAPA_PLANILLA_CUENTAS (backend)"),
        (3, "Workflow",                "Solicitud avanza: Borrador → Objetivos → Presidencia → Directorio",
         "planificacion.solicitud.estado_workflow"),
        (4, "Aprobado Directorio",     "La solicitud queda como presupuesto vigente del ciclo",
         "core.ciclo_presupuestario.estado = 'vigente'"),
        (5, "Export a K2B",            "Cada línea exporta los 5 campos K2B usando item.k2b_item_id, cuenta.k2b_cuenta_id, plan_codigo, monto, periodo",
         "integracion_k2b.export_run (pendiente)"),
        (6, "Ejecución importada",     "K2B genera movimientos (Compromiso → Devengado → Pagado) que vuelven al sistema",
         "ejecucion.movimiento (tipos: PRESUPORDENCOMPRA, PRESUPFACTCONREF, etc.)"),
        (7, "Comparación / Análisis",  "Aprobado vs Ejecutado por cuenta, VP, categoría → cuadros DPP",
         "vistas materializadas en schema analisis"),
    ]
    for fila in flujo:
        ws5.append(list(fila))
    for i, w in enumerate([4, 24, 70, 50], start=1):
        ws5.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="mapa-relaciones.xlsx"'},
    )
